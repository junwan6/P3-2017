#!/usr/bin/env python

#
#  export_database.py
#
# This script reads in spreadsheets to output a batch importable file for the
# Occupation relation.
#
# Future work here: rename this script, because it's rather misleading.

import openpyxl
import os.path
import sys
import warnings

# Validate the input arguments
if len(sys.argv) != 4:
    sys.exit("export_database: missing filenames; first filename is the input file for national data, second filename is the input file for regional data, third filename is the output file")
if not os.path.isfile(sys.argv[1]):
    sys.exit("export_database: {0} not found".format(sys.argv[1]))
if not os.path.isfile(sys.argv[2]):
    sys.exit("export_database: {0} not found".format(sys.argv[2]))
if os.path.dirname(sys.argv[3]) != "" and not os.path.exists(os.path.dirname(sys.argv[2])):
    sys.exit("export_database: directory {0} does not exist".format(sys.argv[2]))

def wageOutOfRange(decimal):
    return unicode(decimal) == u"#"

def formatDecimal(decimal):
    return unicode(decimal).replace(u",", u"").replace(u"$", u"").replace(u">=", u"")

educationDict = { u"No formal education credential" : u"none",
                  u"High school diploma or equivalent" : u"high school",
                  u"Postsecondary nondegree award" : u"postsecondary nondegree",
                  u"Associate's degree" : u"associate",
                  u"Bachelor's degree" : u"bachelor",
                  u"Master's degree" : u"master",
                  u"Doctoral or professional degree" : u"doctoral or professional" }

# First open the national data and store the relevant information in memory;
# we do this because it is the smaller of the two datasets

# Due to the names of the sheets in the spreadsheet, a warning is raised about
# a sheet name conflicting with a reserved name. Fortunately, this does not
# affect the sheet that we're interested in, so suppress this warning.
with warnings.catch_warnings():
    warnings.simplefilter('ignore')
    workbook = openpyxl.load_workbook(sys.argv[1], read_only=True)

try:
    worksheet = workbook.get_sheet_by_name("Table 1.7")
except KeyError:
    sys.exit("export_database: worksheet \"Table 1.7\" not found, please check to make sure you are using the Bureau Labor of Statistics 2014 occupational data")

class PartialOccupation:
    def __init__(self, title, currentEmployment, futureEmployment, careerGrowth, jobOpenings, educationRequired):
        self.title = title
        self.currentEmployment = currentEmployment
        self.futureEmployment = futureEmployment
        self.careerGrowth = careerGrowth
        self.jobOpenings = jobOpenings
        self.educationRequired = educationRequired

partialOccupationData = {}

# Read all rows except the header row
rowCount = 0
for row in worksheet.rows:
    rowCount += 1

    # Skip the first three rows
    if rowCount <= 3:
        continue

    # Ignore row if it is not the data for a single occupation
    if row[2].value is None or row[2].value.strip() != u"Line item":
        continue

    try:
        educationRequired = educationDict[row[10].value.strip()]
    except KeyError:
        educationRequired = u"none"

    # For now, we are not considering jobs that require less than a Bachelor's
    if (educationRequired == u"none" or
        educationRequired == u"high school" or
        educationRequired == u"postsecondary nondegree" or
        educationRequired == u"associate"):
        continue

    partialOccupationData[row[1].value.strip()] = PartialOccupation(row[0].value, formatDecimal(row[3].value), formatDecimal(row[4].value), formatDecimal(row[6].value), formatDecimal(row[8].value), educationRequired)

# Now open the "regional" data and acquire the relevant national data
workbook = openpyxl.load_workbook(sys.argv[2], read_only=True)

# Unfortunately, it seems that each dataset provided by the BLS hardcodes
# the month and year of the data into the sheet, so there is no simple way
# to generalize this sheet name.
sheetName = "All May 2015 Data"
try:
    worksheet = workbook.get_sheet_by_name(sheetName)
except KeyError:
    sys.exit("export_regional_database: worksheet \"{0}\" not found".format(sheetName))

# Pre-compute the "milestone rows," which indicate when we should print out
# a progress report
milestoneRows = [worksheet.max_row * i / 10 for i in range(1, 11)]

# Stream each row, join it with our partial occupation data, and then write it out
with open(sys.argv[3], "w") as outfile:
    rowCount = 0
    for row in worksheet.rows:
        rowCount += 1

        # Print out a status message every 10%
        if rowCount in milestoneRows:
            progressPercent = (milestoneRows.index(rowCount) + 1) * 10
            print 'Export progress: {0}%'.format(progressPercent)

        # Skip the first row
        if rowCount <= 1:
            continue

        # Ignore rows that are not country-wide data
        if row[2].value != u"1":
            continue
        # Ignore rows that are for specific industries
        if row[3].value != u"000000":
            continue
        # Ignore row for aggregations of occupations
        if row[8].value.strip() != u"detail":
            continue

        soc = row[6].value
        # Skip rows that don't have partial occupation data for it
        if not soc in partialOccupationData:
            continue

        occupation = partialOccupationData[soc]

        # Wages specified with a # indicate that the annual wage is at least $187,200, or that the hourly wage is at least $90
        if not (row[28].value is None) and row[28].value == 1:
            # Hourly wage
            wageType = 'hourly'

            averageWage = u"90" if wageOutOfRange(row[14].value) else formatDecimal(row[14].value)
            averageWageOutOfRange = u"1" if wageOutOfRange(row[14].value) else u"0"

            lowWage = u"90" if wageOutOfRange(row[17].value) else formatDecimal(row[17].value)
            lowWageOutOfRange = u"1" if wageOutOfRange(row[17].value) else u"0"

            medianWage = u"90" if wageOutOfRange(row[19].value) else formatDecimal(row[19].value)
            medianWageOutOfRange = u"1" if wageOutOfRange(row[19].value) else u"0"

            highWage = u"90" if wageOutOfRange(row[21].value) else formatDecimal(row[21].value)
            highWageOutOfRange = u"1" if wageOutOfRange(row[21].value) else u"0"
        else:
            # Annual wage
            wageType = 'annual'

            averageWage = u"187200" if wageOutOfRange(row[15].value) else formatDecimal(row[15].value)
            averageWageOutOfRange = u"1" if wageOutOfRange(row[15].value) else u"0"

            lowWage = u"187200" if wageOutOfRange(row[22].value) else formatDecimal(row[22].value)
            lowWageOutOfRange = u"1" if wageOutOfRange(row[22].value) else u"0"

            medianWage = u"187200" if wageOutOfRange(row[24].value) else formatDecimal(row[24].value)
            medianWageOutOfRange = u"1" if wageOutOfRange(row[24].value) else u"0"

            highWage = u"187200" if wageOutOfRange(row[26].value) else formatDecimal(row[26].value)
            highWageOutOfRange = u"1" if wageOutOfRange(row[26].value) else u"0"

        outfile.write(u"\t".join([soc, occupation.title, wageType, averageWage, averageWageOutOfRange, lowWage, lowWageOutOfRange, medianWage, medianWageOutOfRange, highWage, highWageOutOfRange, occupation.educationRequired, occupation.currentEmployment, occupation.futureEmployment, occupation.careerGrowth, occupation.jobOpenings]).encode("UTF-8"))
        outfile.write(u"\n")
